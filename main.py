import numpy as np
import string
import preprocessing as pre
import tfidf as weighting
import openpyxl
import re
import naiveBayes as naive

def read_file(filename):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active 
    m_row = sheet_obj.max_row 
    temp = [] 
    for i in range(1, m_row + 1): 
        cell_obj = sheet_obj.cell(row = i, column = 1) 
        temp.append(cell_obj.value.split("\n"))
    return temp

def clean(document):
    doc = []
    translator = str.maketrans('','',string.punctuation)
    for i in range(len(document)):
        temp = []
        for j in range(len(document[i])):
            document[i][j] = document[i][j].translate(translator)
            document[i][j] = re.sub(r'\d+', '', document[i][j])
            temp.append(document[i][j])
        doc.append(temp)
    return doc

def getValue(document):
    value =[]
    for i in range(len(document)):
        for j in range(len(document[i])):
            value.append(document[i][j][-1:])
    return value

f = open('stopword.txt', 'r')
content = f.read()
spl = content.rstrip()
stop = spl.split()

read = read_file("komentar.xlsx")
#print(read)
valueClass = getValue(read)
kelas = list(map(int,valueClass))
print(kelas)

# Hapus Tanda Baca
cleaner = clean(read)
# print('Cleaner : ',cleaner)
# print("\n")
#Filtering (menghapus kata tidak penting)
tokenization = pre.tokenization(cleaner)
# print('Tokenization',tokenization)
filtering = pre.filtering(tokenization,stop)
removing = pre.remove(filtering)
# print('Removing : ',removing)
# print("\n")
#Stemming (Jadi kata dasar)
stemming = pre.stemming(removing)
# print('Stemming :',stemming)
# print("\n")
#Term hasil preprocessing
term = pre.term(stemming)
count_term = len(term)
# print('Term :',term)

#TF IDF
binary = weighting.binaryWeighting(stemming,term)
# print("\nBinary Weighting:",binary)
raw = weighting.rawWeighting(stemming,term)
# print("\nRaw Weighting:",raw)
gabung = naive.gabung(kelas,raw)
#zip gabung yg lain biar bisa jalan line 79 - 99 (valueKelas1 - sumValueKelas0)
gabung1 = zip(kelas,raw)

#Memisahkan antara kelas 1 & kelas 0
valueKelas1=[]
valueKelas0=[]
for i in gabung1:
    if i[0] == 0:
        valueKelas0.append(i[1])
    elif i[0] == 1:
        valueKelas1.append(i[1])

#Jumlah Semua Value Kelas 1 Index
sumValueKelas1=[]
for i in range(0,len(valueKelas1[0])):
    countValueKelas1=0
    for j in range(0,len(valueKelas1)):
        countValueKelas1+=valueKelas1[j][i]
    sumValueKelas1.append(countValueKelas1)

#Jumlah Semua Value Kelas 0 per Index
sumValueKelas0=[]
for i in range(0,len(valueKelas0[0])):
    countValueKelas0=0
    for j in range(0,len(valueKelas0)):
        countValueKelas0+=valueKelas0[j][i]
    sumValueKelas0.append(countValueKelas0)

# print(list(gabung))
count_dokumen = []
for tup in gabung:
    count_dokumen.append(sum(tup[1]))
count_setiap_dokumen = zip(kelas,count_dokumen)
countPerKelas = naive.countKelas(count_setiap_dokumen)
print(countPerKelas)

peluang0 = naive.peluang0(sumValueKelas0,countPerKelas,count_term)
peluang1 = naive.peluang0(sumValueKelas1,countPerKelas,count_term)
print('peluang 0',peluang0)
print('peluang 1',peluang1)


prior = naive.prior(kelas)
print(prior)




dataT = read_file("datatraining.xlsx")
cleanerData = clean(dataT)
token = pre.tokenization(cleanerData)
filterdata = pre.filtering(token,stop)
reomvedata = pre.remove(filterdata)
stemmingData = pre.stemming(reomvedata)
rawData = weighting.rawWeighting(stemmingData,term)
print(rawData)
likelihood = naive.likelihood(rawData,peluang0,peluang1,prior)
print(likelihood)

# print(list(count_setiap_dokumen))

    # for i in range (len(raw)):
    #     if kelas==0:
    #         count.append(sum(raw[i]))



# log = weighting.logFrequency(raw)
# print("\nLog Frequency:",log)
# DF = weighting.docFrequency(binary)
# print("\nDocument Frequency:",DF)
# IDF = weighting.idf(DF,cleaner)
# print("\nInverse Document Frequency :",IDF)
# hasil_tfidf = weighting.TFIDF(log, IDF)
# print("\ntf-idf:",hasil_tfidf)



